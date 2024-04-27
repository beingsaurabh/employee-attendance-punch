import tkinter as tk
from tkinter import ttk, messagebox, PhotoImage
from datetime import datetime, timedelta
import openpyxl
from openpyxl import load_workbook
from openpyxl.utils.dataframe import dataframe_to_rows
import pandas as pd
import random
from openpyxl.styles import Font
import io
import msoffcrypto
import openpyxl
import os
import subprocess
from pathlib import Path
from openpyxl import Workbook, load_workbook
import time
from fpdf import FPDF

def update_clock():
    current_time = time.strftime("%H:%M:%S")
    current_date = time.strftime("%d-%m-%Y")
    current_day = time.strftime("%A")

    # Convert the date to the desired format
    day, month, year = map(int, current_date.split('-'))
    ordinal_indicator = "th" if 11 <= day <= 13 else {1: "st", 2: "nd", 3: "rd"}.get(day % 10, "th")
    current_date = f"{day}{ordinal_indicator} {time.strftime('%b')}, {year}"

    clock_label.config(text=f"{current_time}, {current_date}, {current_day}")
    clock_label.after(1000, update_clock)  # Update every 1000 milliseconds (1 second)

def count_sundays_between_dates(start_date, end_date):
    try:
        # Convert start_date and end_date to datetime objects
        start_datetime = datetime.strptime(start_date, "%Y-%m-%d %H:%M:%S")
        end_datetime = datetime.strptime(end_date, "%Y-%m-%d %H:%M:%S")
        
        # Initialize a counter for Sundays
        sunday_count = 0
        
        # Iterate through the dates between start_date and end_date
        current_date = start_datetime
        while current_date <= end_datetime:
            # Check if the current date is a Sunday
            if current_date.weekday() == 6:  # Sunday is represented by 6 in Python
                sunday_count += 1
            # Move to the next day
            current_date += timedelta(days=1)
        
        return sunday_count
    except:
        return 0

def calculate_duration(in_time, out_time):
    if in_time is None or out_time is None:
        return None
    in_datetime = datetime.strptime(in_time, "%Y-%m-%d %H:%M:%S")
    out_datetime = datetime.strptime(out_time, "%Y-%m-%d %H:%M:%S")
    duration = out_datetime - in_datetime
    return duration

def round_time(time_str, action):
    time_object = pd.to_datetime(time_str)
    if action == "In":
        rounded_time = time_object.ceil("15T")
    elif action == "Out":
        rounded_time = time_object.floor("15T")
    return rounded_time.strftime("%Y-%m-%d %H:%M:%S")

def get_day_name(date_string):
    date_object = datetime.strptime(date_string, "%Y-%m-%d %H:%M:%S")
    return date_object.strftime("%A")

def authentication_successful():
    label_authentication.config(text="\u2713 User authenticated", foreground="green" )

def punch_time(action):
    employee_name = combo_employee_name.get()
    passcode = entry_passcode.get()

    excel_file = f"DB//session_{employee_name}.xlsx"
    excel_file = os.path.join(os.getcwd(), excel_file) 
    password = "123"      

    # Load passcode from DB//DB-do-not-open.xlsx
    try:
        temp = io.BytesIO()
 
        with open('DB//DB-do-not-open.xlsx', 'rb') as f:
            excel = msoffcrypto.OfficeFile(f)
            excel.load_key('123')
            excel.decrypt(temp)
    
        df = pd.read_excel(temp, sheet_name='DB-do-not-open')        
        del temp  

        passcode_matched = str(df.loc[df['name'] == employee_name, 'passcode'].values[0])

        if passcode == passcode_matched:
            authentication_successful()
            workbook = create_or_load_punch_sheet(employee_name)
            sheet = workbook.active

            # Check if the previous session has an "Out" time recorded when "In" button is clicked again
            if action == "In" and sheet.max_row > 1:
                prev_out_time = sheet.cell(row=sheet.max_row - 1, column=2).value
                if pd.isnull(prev_out_time):
                    prev_in_time = sheet.cell(row=sheet.max_row - 2, column=2).value
                    messagebox.showwarning("Warning", f"Terminating previous session with 'Out' time same as previous 'In' time: {prev_in_time}")
                    rounded_prev_in_time = round_time(prev_in_time, "In")
                    sheet.cell(row=sheet.max_row - 1, column=2, value=rounded_prev_in_time)
                    sheet.cell(row=sheet.max_row - 1, column=3, value=get_day_name(rounded_prev_in_time))
                    sheet.cell(row=sheet.max_row - 1, column=4, value="OUT")
                elif sheet.cell(row=sheet.max_row, column=4).value == "IN":
                    messagebox.showwarning("Warning", "Cannot punch 'In' again without 'Out' entry.")
                    return
            elif action == "Out" and sheet.max_row > 0:
                if sheet.cell(row=sheet.max_row, column=4).value == "OUT":
                    messagebox.showwarning("Warning", "Cannot punch 'Out' again without 'In' entry.")
                    return

            # Process action
            if action == "In":
                session_no = (sheet.max_row // 2) + 1
                rounded_time = round_time(datetime.now().strftime("%Y-%m-%d %H:%M:%S"), "In")
                sheet.append([f"session_{session_no}", rounded_time, get_day_name(rounded_time), "IN", ""])
            elif action == "Out":
                if pd.isnull(sheet.cell(row=sheet.max_row, column=2).value):
                    messagebox.showerror("Error", "Cannot punch 'Out' without 'In' entry.")
                    return
                rounded_time = round_time(datetime.now().strftime("%Y-%m-%d %H:%M:%S"), "Out")
                sheet.append(["", rounded_time, get_day_name(rounded_time), "OUT", ""])
                # Calculate and update duration
                in_time = sheet.cell(row=sheet.max_row - 1, column=2).value
                out_time = sheet.cell(row=sheet.max_row, column=2).value
                duration = calculate_duration(in_time, out_time)
                duration -= timedelta(minutes=30)  # Deduct 30 minutes   

                hours = duration.seconds // 3600
                minutes = (duration.seconds // 60) % 60
                seconds = duration.seconds % 60           

                if int(duration.total_seconds())<0:
                    duration = pd.to_datetime("00:00:00")
                    duration = duration.strftime("%H:%M:%S")                    
                    sheet.cell(row=sheet.max_row - 1, column=5, value=str(duration))
                elif (int(duration.total_seconds())//3600)>23:
                    duration = pd.to_datetime("08:00:00")
                    duration = duration.strftime("%H:%M:%S")                    
                    sheet.cell(row=sheet.max_row - 1, column=5, value=str(duration))
                    messagebox.showwarning("Caution! Unusual Exit", f"Hey {employee_name}!, \n\nDuration of duty cannot be more than 24 hours. Looks like you forgot to mark OUT last time. Hence, Marking as general duty of 8 hours. Be mindful from next time.")
                else:
                    formatted_time = "{:02}:{:02}:{:02}".format(hours, minutes, seconds)
                    duration = pd.to_datetime(formatted_time)
                    duration = duration.strftime("%H:%M:%S")  
                    sheet.cell(row=sheet.max_row - 1, column=5, value=str(duration))

            workbook.save(f"DB//session_{employee_name}.xlsx")
            messagebox.showinfo("Success", f"Time punched {action} successfully!")
            set_password(excel_file, password)            
            show_punch_sheet()
            return
        else:
            label_authentication.config(text=" \u274C Invalid passcode",font=("Ubuntu", 13), foreground="red" )
            messagebox.showerror("Error", "Invalid passcode. Contact Admin if you forgot passcode.")
    except Exception as e:
        messagebox.showerror("Error", f"Some went wrong. Contact Admin (9990158595) {e}")

def set_password(excel_file_path, pw):

    excel_file_path = Path(excel_file_path)
    vbs_script = \
    f"""' Save with password required upon opening
    Set excel_object = CreateObject("Excel.Application")
    Set workbook = excel_object.Workbooks.Open("{excel_file_path}")
    excel_object.DisplayAlerts = False
    excel_object.Visible = False
    workbook.SaveAs "{excel_file_path}",, "{pw}"
    excel_object.Application.Quit
    """

    # write
    vbs_script_path = excel_file_path.parent.joinpath("set_pw.vbs")
    with open(vbs_script_path, "w") as file:
        file.write(vbs_script)
    #execute
    # subprocess.call(['cscript.exe', str(vbs_script_path)])
    # Execute the VBScript using subprocess
    subprocess.run(["cscript.exe", "//NoLogo", str(vbs_script_path)], stdout=subprocess.PIPE, stderr=subprocess.PIPE, creationflags=subprocess.CREATE_NO_WINDOW)

    # remove
    vbs_script_path.unlink()
    return None

def create_or_load_punch_sheet(employee_name):  
    excel_file = f"DB//session_{employee_name}.xlsx"
    excel_file = os.path.join(os.getcwd(), excel_file) 
    password = "123"      

    try:      
        temp = io.BytesIO()

        with open(excel_file, 'rb') as f:
                excel = msoffcrypto.OfficeFile(f)
                excel.load_key('123')
                excel.decrypt(temp)

        workbook = load_workbook(io.BytesIO(temp.getvalue()))
        del temp
        return workbook
    except:
        set_password(excel_file, password)
        workbook = load_workbook(excel_file) 
        return workbook 

def update_punch_sheet_frame(employee_name):

    label_current_display_record.config(text=f"Showing {employee_name}'s record currently", font=("Ubuntu", 8), foreground="purple" )

    workbook = create_or_load_punch_sheet(employee_name)
    sheet = workbook.active

    # Create a pandas DataFrame from the Excel sheet data
    data = []
    for row in sheet.iter_rows(values_only=True):
        data.append(row)
    headers = data.pop(0)
    df = pd.DataFrame(data, columns=headers)
    # Assuming df is your DataFrame
    # df.iloc[:, 0].fillna(' ', inplace=True)
    df.fillna(' ', inplace=True)

    # Clear existing punch sheet frame
    for widget in frame_punch_sheet.winfo_children():
        widget.destroy()

    # Create a treeview widget to display punch data
    tree = ttk.Treeview(frame_punch_sheet, columns=headers, show="headings")
    tree.column("#1", width=90)
    tree.column("#2", width=150)
    tree.column("#3", width=120)
    tree.column("#4", width=120)
    tree.column("#5", width=120)
    for col in headers:
        tree.heading(col, text=col)
    for row in dataframe_to_rows(df, index=False, header=False):
        tree.insert("", "end", values=row)
    tree.pack(side="left", fill="x", expand=False)

    # Add a scrollbar
    scroll_y = ttk.Scrollbar(frame_punch_sheet, orient="vertical", command=tree.yview)
    scroll_y.pack(side="left", fill="y")
    tree.configure(yscrollcommand=scroll_y.set)
    # Adjust scrollbar to show latest data
    tree.yview_moveto(1.0)  # Move scrollbar to the bottom

def get_screen_dimensions():
    # Get the screen width and height
    screen_width = root.winfo_screenwidth()
    screen_height = root.winfo_screenheight()
    return screen_width, screen_height

def restart_app():
    root.destroy()
    main()

def show_punch_sheet():
    update_punch_sheet_frame(combo_employee_name.get())
    frame_punch_sheet.pack(side="left", padx=10, pady=2)

def open_info():
    messagebox.showinfo("Instructions","\u2022 Punch IN and Punch OUT time will be the next & previous multiple of 15 mins of current hour respectively.\n\n\u2022 The company shall not be held liable for any financial losses incurred by employees resulting from negligence in recording entry and exit times.\n\n\u2022 Do not disclose your passcode to anybody.\n\n\u2022 Any effort to tamper with data will be regarded very seriously and addressed in accordance with the company's policy on unfair practices.")

def sum_time_strings(timestrings):
    total_hours = 0
    total_minutes = 0

    for timestr in timestrings:
        hours, minutes, seconds = map(int, timestr.split(":"))
        total_hours += hours
        total_minutes += minutes

    total_hours += total_minutes // 60
    total_minutes = total_minutes % 60

    return total_hours, total_minutes

def gen_non_existing_sheets_with_pwd(employees_list):
    for employee_name in employees_list:
        excel_file = f"DB//session_{employee_name}.xlsx"
        excel_file = os.path.join(os.getcwd(), excel_file)
        password = "123"
        if not os.path.exists("Export_Session_Sheets"):
            os.makedirs("Export_Session_Sheets")

        if not os.path.exists(excel_file):            
            # Create a new workbook
            workbook = openpyxl.Workbook()
            sheet = workbook.active
            sheet.append(["Session_ID", "TimeStamp", "Day", "Nature of Entry", "Work Duration"])
            # Make headers bold
            for cell in sheet["1:1"]:
                cell.font = Font(bold=True)

            workbook.save(excel_file)
            workbook.close()
            set_password(excel_file, password)

class PDF(FPDF):
    def footer(self):
        current_time = time.strftime("%H:%M:%S")
        current_date = time.strftime("%d/%m/%Y")
        current_day = time.strftime("%A")
        self.set_y(-15)
        self.set_font("Arial", style="I", size=8)
        self.cell(0, 10, f"Confidential Document. To be handled only by authorised personnel.     |     Printed on {current_time}, {current_date}, {current_day}", align="C")

def export_session_details():
    employee_name = combo_employee_name.get()
    from_session=entry_from_session.get()
    to_session=entry_to_session.get()
    try:
        if int(from_session)<=int(to_session) and int(from_session)>0 and int(to_session)>0:
            workbook = create_or_load_punch_sheet(employee_name)
            sheet = workbook.active

            # Create a pandas DataFrame from the Excel sheet data
            data = []
            for row in sheet.iter_rows(values_only=True):
                data.append(row)
            headers = data.pop(0)
            df = pd.DataFrame(data, columns=headers)
            df.fillna('', inplace=True)

            # Convert from_session and to_session to integers
            from_session = int(from_session)
            to_session = int(to_session)

            # Filter the DataFrame based on from_session and to_session
            new_df = df.iloc[(from_session-1)*2:to_session*2]
            
            time_strings = []   
            
            # Convert "Work Duration" column to timedelta
            for time in new_df["Work Duration"]:
                if time!="":
                    time_strings.append(str(time))        
            
            total_hours, total_minutes = sum_time_strings(time_strings)

            first_value_of_new_df = new_df["TimeStamp"].iloc[0]
            last_value_of_new_df = new_df["TimeStamp"].iloc[-1]

            total_no_of_sundays = count_sundays_between_dates(first_value_of_new_df, last_value_of_new_df)

            # New row data
            new_row = {
                "Session_ID": ["Total Hours"],
                "TimeStamp": [""],
                "Day": [""],
                "Nature of Entry": [""],
                "Work Duration": ["{} hours {} mins".format(total_hours, total_minutes)]
                }

            df2 = pd.DataFrame(new_row)
            new_df = pd.concat([new_df, df2], ignore_index=True)

            sunday_time_strings = []

            for i in range(total_no_of_sundays):
                sunday_time_strings.append("08:00:00")
                time_strings.append("08:00:00") 

            sunday_total_hours, sunday_total_minutes = sum_time_strings(sunday_time_strings)

            # New row data
            new_row = {
                "Session_ID": ["Sunday Holiday Hours"],
                "TimeStamp": [""],
                "Day": [""],
                "Nature of Entry": [""],
                "Work Duration": ["{} hours {} mins".format(sunday_total_hours, sunday_total_minutes)]
                }

            df2 = pd.DataFrame(new_row)
            new_df = pd.concat([new_df, df2], ignore_index=True)
            
            total_hours, total_minutes = sum_time_strings(time_strings)

            # New row data
            new_row = {
                "Session_ID": ["Grand Total Hours"],
                "TimeStamp": [""],
                "Day": [""],
                "Nature of Entry": [""],
                "Work Duration": ["{} hours {} mins".format(total_hours, total_minutes)]
                }

            df3 = pd.DataFrame(new_row)
            new_df = pd.concat([new_df, df3], ignore_index=True)

            filename = f"Export_Session_Sheets/{employee_name}_session_{from_session}_to_{to_session}.xlsx"
            pdffile = f"Export_Session_Sheets/{employee_name}_session_{from_session}_to_{to_session}.pdf"       
            new_df.to_excel(filename, index=False)

            # Open the Excel file
            workbook = load_workbook(filename)
            worksheet = workbook.active

            # Determine the total number of rows in the worksheet
            total_rows = len(worksheet['A'])

            # Bold the last three rows
            for row_idx in range(total_rows - 2, total_rows + 1):
                for cell in worksheet[row_idx]:
                    cell.font = Font(bold=True)            

            # Auto-adjust the width of all columns
            for column_cells in worksheet.columns:
                max_length = 0
                for cell in column_cells:
                    try:
                        if len(str(cell.value)) > max_length:
                            max_length = len(cell.value)
                    except:
                        pass
                adjusted_width = (max_length + 2) * 1.2  # Adjusted width (add some extra padding)
                worksheet.column_dimensions[cell.column_letter].width = adjusted_width

            # Save the changes
            workbook.save(filename)   
            # Create PDF instance
            pdf = PDF()
            pdf.set_auto_page_break(auto=True, margin=10)
            pdf.add_page()
            # Add heading to PDF
            pdf.set_font("Arial",style="B", size=15)
            pdf.cell(200, 10, txt=f"Toolmaster Industries", ln=True, align="C")
            pdf.set_font("Arial",style="B", size=13)
            pdf.cell(200, 10, txt=f"Employee Time Sheet", ln=True, align="C")
            pdf.set_font("Arial",style="B", size=11)
            pdf.cell(200, 10, txt=f"Employee Name: {employee_name}  |   From Session: {from_session}   |    To Session: {to_session}", ln=True, align="C")
            pdf.set_font("Arial", size=9)

            # Iterate over rows and columns in the worksheet
            for i, row in enumerate(worksheet.iter_rows()):
                for j, cell in enumerate(row):
                    # Check if cell value is None (blank)
                    if cell.value is None:
                        cell_value = ""
                    else:
                        cell_value = str(cell.value)

                    # Set fill color for headers (first row)
                    if i == 0:
                        pdf.set_fill_color(192, 192, 192)  # Light gray
                    # Set fill color for last three rows
                    elif total_rows - i - 1 <= 2:
                        pdf.set_fill_color(255, 204, 204)  # Light red for last three rows
                    else:
                        pdf.set_fill_color(255, 255, 255)  # White
                    
                    # Draw cell with background color
                    pdf.cell(37, 5, cell_value, border=1, fill=True)
                
                pdf.ln()       
            
            # Save PDF to file
            pdf.output(pdffile)  
            messagebox.showinfo("Exported Successfully", f"Saved as {filename}")
        else:
            messagebox.showerror("Error", "'From session value' can't be greater than 'to session value Or Negative'")
    except Exception as e:
        messagebox.showerror("Invalid Input", f"Input session ID is not valid {e}")
           
def main():
    global root, frame_punch_sheet, combo_employee_name, entry_passcode, label_authentication, entry_from_session, entry_to_session, label_current_display_record, clock_label    

    root = tk.Tk()
    root.title("Employee Time Punch App - Toolmaster Industries")
    
    # Create a frame for the navigation bar
    nav_frame = tk.Frame(root, bg="white")
    nav_frame.pack(side="top", fill="x")

    # Create navigation buttons
    restart_button = tk.Button(nav_frame, text="\u21BB Restart", command=restart_app)
    restart_button.pack(side="right", padx=2, pady=2)

    home_button = tk.Button(nav_frame, text="\u2139 Info", command=open_info)
    home_button.pack(side="right", padx=6, pady=2)

    label_company_name = tk.Label(nav_frame, text="Toolmaster Industries - Self Time Punch Portal",font=("Ubuntu", 16), fg="purple", bg="white")
    label_company_name.pack(side="left", padx=2, pady=2, fill="both", expand=True)

    # Get the screen dimensions
    screen_width, screen_height = get_screen_dimensions()

    # Calculate the desired width and height (3/4th of the screen dimensions)
    win_width = int(screen_width * 3 / 4)
    win_height = int(screen_height * 9 / 10)

    # Set the window size
    root.geometry(f"{win_width}x{win_height}")
    root.state("zoomed")

    # Load employee names from the Excel file
    try:
        temp = io.BytesIO()
 
        with open('DB//DB-do-not-open.xlsx', 'rb') as f:
            excel = msoffcrypto.OfficeFile(f)
            excel.load_key('123')
            excel.decrypt(temp)
    
        df = pd.read_excel(temp, sheet_name='DB-do-not-open')        
        del temp  

        employee_names = df['name'].tolist()
    except Exception as e:
        messagebox.showerror("Error", e)
        employee_names = []
    
    # Load the image
    original_image = PhotoImage(file="DB//Assets//Logo.png")

    # Resize the image
    width, height = original_image.width(), original_image.height()
    resized_image = original_image.subsample(max(1, width//120), max(1, height//50))

    frame0 = tk.Frame(root)
    frame0.pack(pady=2)

    # Create a label to display the resized image
    image_label = tk.Label(frame0, image=resized_image)
    image_label.pack(side="left", padx=30, pady=2)

    frame1 = tk.Frame(root)
    frame1.pack(pady=2)
    
    # Create labels and entry widgets
    label_employee_name = tk.Label(frame1, text="Employee Name :",font=("Ubuntu", 13))
    label_employee_name.pack(side="left", padx=10, pady=2)
    combo_employee_name = ttk.Combobox(frame1, values=sorted(employee_names),font=("Ubuntu", 13), state="readonly")
    combo_employee_name.current(random.randrange(len(employee_names)))
    combo_employee_name.pack(side="left", padx=10, pady=2)

    frame2 = tk.Frame(root)
    frame2.pack(pady=2)

    label_passcode = tk.Label(frame2, text="Passcode :",font=("Ubuntu", 13))
    label_passcode.pack(side="left", padx=25, pady=2)
    entry_passcode = tk.Entry(frame2, show="*",font=("Ubuntu", 13))
    entry_passcode.pack(side="left", padx=10, pady=2)

    frame7 = tk.Frame(root)
    frame7.pack(pady=2)

    label_authentication = tk.Label(frame7, text=" \u26A0 User not authenticated",font=("Ubuntu", 13), foreground="brown")
    label_authentication.pack(side="top", padx=5, pady=2)

    frame3 = tk.Frame(root)
    frame3.pack(pady=1)

    # Create buttons
    button_in = tk.Button(frame3, text="In \u2199", width=20, command=lambda: punch_time("In"),font=("Ubuntu", 13),bg="yellow")
    button_in.pack(side="left", padx=10, pady=2)

    # Create Show My Data button
    button_show_data = tk.Button(frame3, text="Show My Data \U0001F441", command=show_punch_sheet,font=("Ubuntu", 13), bg="light green")
    button_show_data.pack(side="left", padx=10, pady=2)

    button_out = tk.Button(frame3, text="Out \u2197", width=20, command=lambda: punch_time("Out"),font=("Ubuntu", 13), bg="light blue")
    button_out.pack(side="left", padx=10, pady=2)

    frame5 = tk.Frame(root)
    frame5.pack(pady=2)

    label_current_display_record = tk.Label(frame5, text="")
    label_current_display_record.pack(side="top", padx=1, pady=1)

    # Create a frame to display punch sheet
    frame_punch_sheet = tk.Frame(frame5)
    frame_punch_sheet.pack(side="top", padx=10, pady=2)
    
    gen_non_existing_sheets_with_pwd(employee_names)
    show_punch_sheet()

    frame6 = tk.Frame(root)
    frame6.pack(pady=2)

    export_session_data_label = tk.Label(frame6, text="Export Session Data: ",font=("Ubuntu", 13))
    export_session_data_label.pack(side="left", padx=5, pady=2)

    export_session_fron_label = tk.Label(frame6, text="From",font=("Ubuntu", 13))
    export_session_fron_label.pack(side="left", padx=5, pady=2)

    entry_from_session = tk.Entry(frame6, width=8,font=("Ubuntu", 13))
    entry_from_session.pack(side="left", padx=5, pady=2)

    export_session_fron_label = tk.Label(frame6,  text="To",font=("Ubuntu", 13))
    export_session_fron_label.pack(side="left", padx=5, pady=2)

    entry_to_session = tk.Entry(frame6, width=8,font=("Ubuntu", 13))
    entry_to_session.pack(side="left", padx=5, pady=2)

    button_export_data = tk.Button(frame6, text="Export \u2934", command=export_session_details,font=("Ubuntu", 13), bg="light pink")
    button_export_data.pack(side="left", padx=15, pady=2)

    text_area = tk.Text(root, width=87, height=7,font=("Ubuntu", 9),bg="light yellow")
    text_area.insert(tk.END, "\u2022 Punch IN and Punch OUT time will be the next & previous multiple of 15 mins of current hour respectively.\n\u2022 A short lunch break of 30 mins is automatically deducted from work duration.\n\u2022 The company shall not be held liable for any financial losses incurred by employees resulting from negligence   in recording entry and exit times.\n\u2022 Do not disclose your passcode to anybody.\n\u2022 Any effort to tamper with data will be regarded very seriously and addressed in accordance with the company's   policy on unfair practices.")
    text_area.config(state=tk.DISABLED)
    text_area.pack(padx=1, pady=2)  

    # Create a frame for the navigation bar
    footer = tk.Frame(root, bg="white")

    # Create a label for the clock
    clock_label = tk.Label(footer, text="", font=("Ubuntu", 9), fg="black", bg="white")
    clock_label.pack(side="right", padx=2, pady=2)

    footer.pack(side="bottom", fill="x")
    footer_desc = tk.Label(footer, text="Unauthorised use of this portal is strictly prohibited.   |   \u260E Contact Admin: 9990158595",font=("Ubuntu", 13), bg="white",fg="purple")
    footer_desc.pack(side="left", padx=6, pady=2, expand=True)

    root.iconbitmap("DB//Assets//Logo_1.ico")
    
    root.eval('tk::PlaceWindow . center')
    
    # Start updating the clock
    update_clock()

    root.mainloop()

if __name__=="__main__":
    try:
        main()
    except Exception as e:
        messagebox.showerror("Error", e)
